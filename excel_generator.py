# excel_generator.py – Eksport do Excel (openpyxl)

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config import ABACUS

NAVY_HEX = "0D1B2A"
LIGHT_HEX = "F4F6F8"
GREEN_HEX = "27AE60"
ORANGE_HEX = "F39C12"
RED_HEX = "E74C3C"
WHITE_HEX = "FFFFFF"
BORDER_HEX = "DEE2E6"

RISK_FILL = {
    "niskie": PatternFill("solid", fgColor="EAFAF1"),
    "średnie": PatternFill("solid", fgColor="FEF9E7"),
    "wysokie": PatternFill("solid", fgColor="FDE8E8"),
}
RISK_FONT_COLOR = {
    "niskie": "27AE60",
    "średnie": "E67E22",
    "wysokie": "C0392B",
}

thin = Side(style="thin", color=BORDER_HEX)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr_fill():
    return PatternFill("solid", fgColor=NAVY_HEX)


def _alt_fill():
    return PatternFill("solid", fgColor=LIGHT_HEX)


def generate_aggregate_excel(audyty: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport zbiorczy"

    # ----- Nagłówek raportu -----
    ws.merge_cells("A1:I1")
    ws["A1"] = "AUDYT TECZEK AKT OSOBOWYCH – Raport zbiorczy"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=WHITE_HEX)
    ws["A1"].fill = _hdr_fill()
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"] = (
        f"{ABACUS['nazwa']}  |  {ABACUS['adres']}  |  "
        f"{ABACUS['www']}  |  {ABACUS['email']}  |  tel. {ABACUS['tel']}"
    )
    ws["A2"].font = Font(name="Calibri", size=8, color="888888")
    ws["A2"].fill = PatternFill("solid", fgColor="F8F9FA")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:I3")
    ws["A3"] = f"Wygenerowano: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Liczba pracowników: {len(audyty)}"
    ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="555555")
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 18

    # ----- Nagłówki kolumn -----
    headers = [
        "Firma", "Imię i nazwisko", "Stanowisko", "Status",
        "Braki obowiązkowe", "Braki warunkowe", "Braki dobrowolne",
        "Kompletność (%)", "Poziom ryzyka",
    ]
    col_widths = [22, 22, 22, 12, 14, 14, 14, 14, 14]

    ROW_HDR = 5
    ws.row_dimensions[ROW_HDR].height = 22
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=ROW_HDR, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE_HEX)
        cell.fill = PatternFill("solid", fgColor="1B2D45")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ----- Dane -----
    for row_i, a in enumerate(audyty, start=ROW_HDR + 1):
        risk = a.get("poziom_ryzyka", "niskie")
        rfill = RISK_FILL.get(risk, PatternFill("solid", fgColor=WHITE_HEX))
        row_fill = _alt_fill() if (row_i - ROW_HDR) % 2 == 0 else PatternFill("solid", fgColor=WHITE_HEX)

        values = [
            a.get("firma", "—"),
            f"{a.get('imie', '')} {a.get('nazwisko', '')}",
            a.get("stanowisko", "—"),
            a.get("status_pracownika", "—"),
            a.get("braki_obowiazkowe", 0),
            a.get("braki_warunkowe", 0),
            a.get("braki_dobrowolne", 0),
            a.get("kompletnosc_procent", 100),
            a.get("poziom_ryzyka", "niskie").upper(),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col_idx in (5, 6, 7, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_idx == 9:
                cell.fill = rfill
                cell.font = Font(
                    name="Calibri", bold=True, size=9,
                    color=RISK_FONT_COLOR.get(risk, "000000"),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = row_fill

        ws.row_dimensions[row_i].height = 18

    # ----- Autofilter -----
    ws.auto_filter.ref = f"A{ROW_HDR}:I{ROW_HDR + len(audyty)}"

    # ----- Freeze panes -----
    ws.freeze_panes = f"A{ROW_HDR + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_individual_excel(audyt: dict, gap_list: list) -> bytes:
    """Eksport szczegółowy jednego audytu – lista braków."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Braki – szczegóły"

    # Nagłówek
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Audyt: {audyt.get('imie')} {audyt.get('nazwisko')} | {audyt.get('firma')}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=12, color=WHITE_HEX)
    ws["A1"].fill = _hdr_fill()
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = ["ID", "Sekcja", "Dokument", "Typ braku", "Rekomendacja"]
    col_widths = [6, 10, 45, 14, 30]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE_HEX)
        cell.fill = PatternFill("solid", fgColor="1B2D45")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    STATUS_FILL = {
        "obowiązkowe": PatternFill("solid", fgColor="FDE8E8"),
        "warunkowe": PatternFill("solid", fgColor="FEF9E7"),
        "dobrowolne": PatternFill("solid", fgColor="EAFAF1"),
    }

    for row_i, g in enumerate(gap_list, start=4):
        values = [
            g.get("id", "—"),
            g.get("section", "—"),
            g.get("text", "—"),
            g.get("status", "—"),
            g.get("recommendation", "—"),
        ]
        sfill = STATUS_FILL.get(g.get("status", ""), PatternFill("solid", fgColor=WHITE_HEX))
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = sfill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
